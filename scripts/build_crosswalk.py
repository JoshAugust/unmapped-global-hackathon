#!/usr/bin/env python3
"""
Build SOC ↔ ISCO-08 Crosswalk

Produces:
  data/crosswalks/soc_isco_crosswalk.csv
  data/crosswalks/frey_osborne_isco.csv
  data/crosswalks/crosswalk_summary.json

Strategy:
  1. Try downloading the BLS SOC↔ISCO crosswalk Excel file
  2. If download fails, build crosswalk algorithmically using:
     a. Known SOC major-group → ISCO major-group correspondences
     b. Fuzzy title matching between SOC occupations and ISCO-08 occupations
  3. Join with Frey-Osborne automation probabilities
"""

import csv
import json
import os
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

# --- Paths ---
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data" / "crosswalks"
SON_OF_BRIDGE = PROJECT_DIR.parent / "son-of-a-bridge"
FREY_OSBORNE_PATH = SON_OF_BRIDGE / "data" / "frey_osborne" / "automation_probabilities.csv"
ISCO_PATH = SON_OF_BRIDGE / "data" / "ilo_isco08" / "isco08_occupations.csv"

DATA_DIR.mkdir(parents=True, exist_ok=True)

# --- Known SOC Major Group → ISCO-08 Major Group Mapping ---
# SOC uses XX-XXXX format. First two digits = major group.
# ISCO-08 uses 1-digit major groups (0-9).
# This is the well-established BLS/ILO correspondence.
SOC_TO_ISCO_MAJOR = {
    "11": ["1"],           # Management → Managers
    "13": ["1", "2", "3"], # Business & Financial Operations → Managers/Professionals/Technicians
    "15": ["2", "3"],      # Computer & Mathematical → Professionals/Technicians
    "17": ["2", "3"],      # Architecture & Engineering → Professionals/Technicians
    "19": ["2"],           # Life, Physical, Social Science → Professionals
    "21": ["2", "3"],      # Community & Social Service → Professionals/Technicians
    "23": ["2"],           # Legal → Professionals
    "25": ["2"],           # Education, Training, Library → Professionals
    "27": ["2", "3"],      # Arts, Design, Entertainment, Sports → Professionals/Technicians
    "29": ["2", "3"],      # Healthcare Practitioners → Professionals/Technicians
    "31": ["3", "5"],      # Healthcare Support → Technicians/Service Workers
    "33": ["3", "5"],      # Protective Service → Technicians/Service Workers
    "35": ["5"],           # Food Preparation & Serving → Service Workers
    "37": ["9"],           # Building & Grounds Cleaning → Elementary Occupations
    "39": ["5"],           # Personal Care & Service → Service Workers
    "41": ["5", "3"],      # Sales → Service Workers/Technicians
    "43": ["4"],           # Office & Administrative Support → Clerical Support
    "45": ["6"],           # Farming, Fishing, Forestry → Skilled Agriculture
    "47": ["7"],           # Construction & Extraction → Craft Workers
    "49": ["7"],           # Installation, Maintenance, Repair → Craft Workers
    "51": ["7", "8"],      # Production → Craft/Plant Operators
    "53": ["8", "9"],      # Transportation & Material Moving → Plant Operators/Elementary
    "55": ["0"],           # Military → Armed Forces
}

# --- More detailed SOC minor group → ISCO unit group mappings ---
# These provide finer-grained mappings where the major group is too broad
SOC_MINOR_TO_ISCO = {
    # Management occupations
    "11-10": ["1112", "1114"],  # Top Executives → Senior Officials
    "11-20": ["1221", "1222"],  # Advertising/Marketing/PR Managers
    "11-30": ["1211", "1212", "1213", "1219"],  # Operations Specialties Managers
    "11-90": ["1311", "1312", "1321", "1324", "1330", "1341", "1342", "1343", "1344", "1345", "1346", "1349"],
    
    # Business & Financial
    "13-10": ["2411", "2412", "2413"],  # Business Operations Specialists → Finance Professionals
    "13-20": ["2411", "2412", "2413", "2421", "2422", "2423", "2424", "3313"],  # Financial Specialists
    
    # Computer & Mathematical
    "15-10": ["2511", "2512", "2513", "2514", "2519", "2521", "2522", "2523", "2529"],  # Computer → ICT Professionals
    "15-20": ["2120"],  # Mathematical → Mathematicians
    
    # Architecture & Engineering
    "17-10": ["2161", "2162", "2164"],  # Architects/Surveyors
    "17-20": ["2141", "2142", "2143", "2144", "2145", "2146", "2149"],  # Engineers
    "17-30": ["3118", "3119"],  # Drafters/Technicians
    
    # Life, Physical, Social Science
    "19-10": ["2131", "2132", "2133"],  # Life Scientists → Biologists etc
    "19-20": ["2111", "2112", "2113", "2114"],  # Physical Scientists
    "19-30": ["2631", "2632", "2633", "2634"],  # Social Scientists → Economists etc
    "19-40": ["2114", "2131"],  # Life/Physical Science Technicians
    
    # Community & Social Service
    "21-10": ["2635", "2636", "3412"],  # Counselors/Social Workers
    "21-20": ["2635", "2636"],  # Religious Workers
    
    # Legal
    "23-10": ["2611", "2612", "2619"],  # Lawyers/Judges
    "23-20": ["3411"],  # Legal Support
    
    # Education
    "25-10": ["2310", "2320", "2330"],  # Postsecondary Teachers
    "25-20": ["2341", "2342"],  # Primary/Secondary/Special Ed
    "25-30": ["2352", "2353", "2354", "2355", "2356", "2359"],  # Other Teachers
    "25-40": ["2621", "2622"],  # Librarians/Curators
    
    # Arts, Design, Entertainment
    "27-10": ["2161", "2162", "2163", "2166"],  # Art/Design
    "27-20": ["2641", "2642", "2643"],  # Entertainers/Performers
    "27-30": ["2642", "2651", "2652", "2653", "2654", "2655", "2656", "2659"],  # Media/Communication
    "27-40": ["2651", "2652", "2653", "2654", "2655", "2656", "2659", "3435"],  # Media/Communication cont.
    
    # Healthcare Practitioners
    "29-10": ["2211", "2212"],  # Health Diagnosing/Treating → Physicians
    "29-20": ["2261", "2262", "2263", "2264", "2265", "2266", "2267", "2269"],  # Health Technologists
    "29-90": ["2269"],  # Other Health Practitioners
    
    # Healthcare Support
    "31-10": ["5321", "5322", "5329"],  # Nursing/Home Health Aides
    "31-20": ["3256"],  # Occupational/Physical Therapy Assistants
    "31-90": ["5321", "5329", "3259"],  # Other Healthcare Support
    
    # Protective Service
    "33-10": ["1349", "5411", "5412", "5413", "5414", "5419"],  # Supervisors/Law Enforcement
    "33-20": ["5411", "5412"],  # Fire Fighting
    "33-30": ["5412", "5413", "5414"],  # Law Enforcement
    "33-90": ["5414", "5419"],  # Other Protective
    
    # Food Preparation & Serving
    "35-10": ["5120", "5151", "5152"],  # Supervisors/Cooks/Food Prep
    "35-20": ["5120"],  # Cooks/Food Prep
    "35-30": ["5131", "5132"],  # Food/Beverage Serving
    
    # Building & Grounds Cleaning
    "37-10": ["9112", "9121", "9122", "9129"],  # Building Cleaning
    "37-20": ["6113", "9214"],  # Grounds Maintenance
    
    # Personal Care & Service
    "39-10": ["5141", "5142"],  # Supervisors/Gaming
    "39-20": ["5113"],  # Animal Care/Service
    "39-30": ["5141"],  # Entertainment Attendants
    "39-40": ["5163"],  # Funeral Service
    "39-50": ["5142"],  # Personal Appearance
    "39-90": ["5111", "5112", "5113", "5151", "5152", "5153", "5164", "5165", "5169"],  # Other Personal Care
    
    # Sales
    "41-10": ["3321", "3322", "5221", "5222", "5223"],  # Supervisors/Sales
    "41-20": ["3321", "3322", "3323", "3324"],  # Retail Sales
    "41-30": ["2431", "2432", "2433", "2434"],  # Sales Representatives
    "41-40": ["3339"],  # Other Sales
    "41-90": ["5230", "5242", "5243", "5244", "5245", "5246", "5249"],  # Other Sales
    
    # Office & Administrative Support
    "43-10": ["3341", "3342", "3343", "3344"],  # Supervisors/Office
    "43-20": ["3341"],  # Communications Equipment Operators
    "43-30": ["4110", "4120", "4131", "4132"],  # Financial Clerks
    "43-40": ["4211", "4212", "4213", "4214", "4221", "4222", "4223", "4224", "4225", "4226", "4227"],  # Info/Record Clerks
    "43-50": ["4311", "4312", "4313", "4321", "4322", "4323"],  # Material Recording
    "43-60": ["4110", "4120", "4131", "4132", "4411", "4412", "4413", "4414", "4415", "4416", "4419"],  # Secretaries/Admin
    "43-90": ["4415", "4416", "4419"],  # Other Office
    
    # Farming, Fishing, Forestry
    "45-10": ["6111", "6112", "6113", "6114", "6121", "6122", "6123", "6129", "6130"],  # Farm Workers
    "45-20": ["6221", "6222"],  # Fishing/Hunting
    "45-40": ["6210"],  # Forest/Conservation
    
    # Construction & Extraction
    "47-10": ["7111", "7112", "7113", "7114", "7115", "7119"],  # Construction Trades
    "47-20": ["7111", "7112", "7113", "7114", "7115", "7119", "7121", "7122", "7123", "7124", "7125", "7126", "7127"],
    "47-30": ["7131", "7132", "7133"],  # Helpers
    "47-40": ["7541", "7542", "7543", "7544", "7549"],  # Other Construction
    "47-50": ["8111", "8112", "8113", "8114"],  # Extraction Workers
    
    # Installation, Maintenance, Repair
    "49-10": ["7411", "7412", "7413"],  # Supervisors/Mechanics
    "49-20": ["7231", "7232", "7233", "7234"],  # Vehicle Mechanics
    "49-30": ["7411", "7412", "7413", "7421", "7422"],  # Industrial Mechanics
    "49-90": ["7311", "7312", "7313", "7314", "7315", "7316", "7317", "7318", "7319"],  # Other Maintenance
    
    # Production
    "51-10": ["8121", "8122", "8131"],  # Supervisors/Production
    "51-20": ["7211", "7212", "7213", "7214", "7215"],  # Assemblers/Fabricators
    "51-30": ["7511", "7512", "7513", "7514", "7515", "7516"],  # Food Processing
    "51-40": ["7521", "7522", "7523", "7531", "7532", "7533", "7534", "7535", "7536"],  # Metal/Plastic
    "51-50": ["7311", "7312", "7313", "7314", "7315", "7316", "7317", "7318", "7319"],  # Printing
    "51-60": ["8141", "8142", "8143"],  # Textile/Apparel
    "51-70": ["7521", "7522", "7523"],  # Woodworking
    "51-80": ["8131", "8141", "8142", "8143", "8151", "8152", "8153", "8154", "8155", "8156", "8157", "8159"],  # Plant Operators
    "51-90": ["7543", "8181", "8182", "8183", "8189", "9329"],  # Other Production
    
    # Transportation & Material Moving
    "53-10": ["8311", "8312"],  # Air Transportation
    "53-20": ["8321", "8322"],  # Motor Vehicle Operators
    "53-30": ["8331", "8332", "8341", "8342", "8343", "8344", "8350"],  # Rail/Water Transportation
    "53-40": ["8321", "8322"],  # Other Transportation
    "53-50": ["8344"],  # Material Moving Machine Operators
    "53-60": ["9321", "9329", "9331", "9332", "9333", "9334"],  # Laborers/Material Movers
    "53-70": ["9611", "9612", "9613", "9621", "9622", "9623", "9624", "9629"],  # Helpers
    
    # Military
    "55-10": ["0110"],  # Military Officers
    "55-20": ["0210"],  # Military Enlisted
    "55-30": ["0310"],  # Military Specific
}


def normalize_title(title: str) -> str:
    """Normalize occupation title for matching."""
    t = title.lower().strip()
    # Remove common noise words and punctuation
    t = re.sub(r'[,\.\(\)\[\]\-/]', ' ', t)
    t = re.sub(r'\b(all other|except|and|or|the|of|in|for|a|an)\b', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def title_similarity(a: str, b: str) -> float:
    """Compute similarity between two occupation titles."""
    na, nb = normalize_title(a), normalize_title(b)
    # Direct substring match bonus
    if na in nb or nb in na:
        return 0.85 + 0.15 * SequenceMatcher(None, na, nb).ratio()
    return SequenceMatcher(None, na, nb).ratio()


def load_frey_osborne() -> list[dict]:
    """Load Frey & Osborne automation probabilities."""
    rows = []
    with open(FREY_OSBORNE_PATH, newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Clean occupation name (they have no spaces)
            occ = row['occupation']
            # Insert spaces before capitals (CamelCase → separate words)
            occ_clean = re.sub(r'([a-z])([A-Z])', r'\1 \2', occ)
            occ_clean = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', occ_clean)
            occ_clean = occ_clean.replace(',', ', ')
            rows.append({
                'soc_code': row['soc_code'].strip(),
                'occupation': occ_clean,
                'probability': float(row['probability']),
                'rank': int(row['rank']),
            })
    return rows


def load_isco08() -> dict:
    """Load ISCO-08 occupations. Returns dict keyed by code."""
    isco = {}
    with open(ISCO_PATH, newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            isco[row['code'].strip()] = {
                'code': row['code'].strip(),
                'title': row['title'].strip(),
                'level': int(row['level']),
            }
    return isco


def try_download_bls_crosswalk() -> list[dict] | None:
    """Try to download the BLS SOC↔ISCO crosswalk. Returns None on failure."""
    import urllib.request
    import tempfile
    
    urls = [
        "https://www.bls.gov/soc/ISCO_SOC_Crosswalk.xls",
        "https://www.bls.gov/soc/soc_2018_isco_08_crosswalk.xlsx",
    ]
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Referer': 'https://www.bls.gov/soc/soccrosswalks.htm',
    }
    
    for url in urls:
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=15) as resp:
                if resp.status != 200:
                    continue
                content_type = resp.headers.get('Content-Type', '')
                if 'html' in content_type:
                    print(f"  ✗ {url} returned HTML (likely blocked)")
                    continue
                
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(resp.read())
                    tmp_path = tmp.name
                
                # Try parsing with openpyxl
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(tmp_path)
                    ws = wb.active
                    rows = []
                    headers_row = [str(c.value).strip() if c.value else '' for c in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            rows.append(dict(zip(headers_row, [str(v).strip() if v else '' for v in row])))
                    os.unlink(tmp_path)
                    if len(rows) > 50:
                        print(f"  ✓ Downloaded {url}: {len(rows)} rows")
                        return rows
                except ImportError:
                    print("  ✗ openpyxl not installed, cannot parse Excel")
                    os.unlink(tmp_path)
                except Exception as e:
                    print(f"  ✗ Error parsing {url}: {e}")
                    os.unlink(tmp_path)
        except Exception as e:
            print(f"  ✗ Failed to download {url}: {e}")
    
    return None


def build_algorithmic_crosswalk(frey_osborne: list[dict], isco: dict) -> list[dict]:
    """
    Build SOC↔ISCO crosswalk algorithmically using:
    1. Known major group correspondences
    2. Known minor group → ISCO unit group mappings
    3. Fuzzy title matching within constrained groups
    """
    # Get ISCO level-4 occupations (detailed)
    isco_detailed = {k: v for k, v in isco.items() if v['level'] == 4}
    # Also keep level 3 for broader matching
    isco_l3 = {k: v for k, v in isco.items() if v['level'] == 3}
    
    crosswalk = []
    
    for fo in frey_osborne:
        soc_code = fo['soc_code']
        soc_title = fo['occupation']
        soc_major = soc_code[:2]
        soc_minor = soc_code[:5]  # e.g., "11-10"
        
        # Get candidate ISCO major groups
        isco_majors = SOC_TO_ISCO_MAJOR.get(soc_major, [])
        if not isco_majors:
            continue
        
        # Check if we have a specific minor group mapping
        specific_isco_codes = SOC_MINOR_TO_ISCO.get(soc_minor, [])
        
        best_matches = []
        
        if specific_isco_codes:
            # Use specific minor-to-unit group mappings + title matching
            for isco_code in specific_isco_codes:
                if isco_code in isco_detailed:
                    sim = title_similarity(soc_title, isco_detailed[isco_code]['title'])
                    best_matches.append((isco_code, isco_detailed[isco_code]['title'], sim, 'direct'))
                elif isco_code in isco_l3:
                    # Check all level-4 children
                    for k, v in isco_detailed.items():
                        if k.startswith(isco_code[:3]):
                            sim = title_similarity(soc_title, v['title'])
                            best_matches.append((k, v['title'], sim, 'partial'))
        
        # Also do broad title matching within candidate ISCO major groups
        for code, entry in isco_detailed.items():
            if code[0] in isco_majors:
                sim = title_similarity(soc_title, entry['title'])
                if sim > 0.45:
                    best_matches.append((code, entry['title'], sim, 'approximate'))
        
        # Deduplicate by ISCO code, keeping best similarity
        seen = {}
        for isco_code, isco_title, sim, mtype in best_matches:
            if isco_code not in seen or sim > seen[isco_code][1]:
                # Upgrade mapping type if similarity is high
                if sim > 0.7:
                    mtype = 'direct'
                elif sim > 0.5:
                    mtype = 'partial'
                seen[isco_code] = (isco_title, sim, mtype)
        
        # Take top matches (at least 1, up to 3)
        sorted_matches = sorted(seen.items(), key=lambda x: -x[1][1])
        
        if sorted_matches:
            # Always include the best match; include others if they're close
            best_sim = sorted_matches[0][1][1]
            for isco_code, (isco_title, sim, mtype) in sorted_matches[:3]:
                if sim >= max(0.35, best_sim - 0.2):
                    crosswalk.append({
                        'soc_code': soc_code,
                        'soc_title': soc_title,
                        'isco08_code': isco_code,
                        'isco08_title': isco_title,
                        'mapping_type': mtype,
                    })
        else:
            # Fallback: map to ISCO level-3 group
            for code, entry in isco_l3.items():
                if code[0] in isco_majors:
                    sim = title_similarity(soc_title, entry['title'])
                    if sim > 0.3:
                        crosswalk.append({
                            'soc_code': soc_code,
                            'soc_title': soc_title,
                            'isco08_code': code,
                            'isco08_title': entry['title'],
                            'mapping_type': 'approximate',
                        })
                        break
            else:
                # Last resort: map to ISCO major group
                mg = isco_majors[0]
                if mg in isco:
                    crosswalk.append({
                        'soc_code': soc_code,
                        'soc_title': soc_title,
                        'isco08_code': mg,
                        'isco08_title': isco[mg]['title'],
                        'mapping_type': 'approximate',
                    })
    
    return crosswalk


def build_frey_osborne_isco(crosswalk: list[dict], frey_osborne: list[dict]) -> list[dict]:
    """Map Frey-Osborne automation probabilities to ISCO codes."""
    # Build SOC → probability lookup
    soc_probs = {fo['soc_code']: fo['probability'] for fo in frey_osborne}
    
    # Group by ISCO code
    isco_groups = defaultdict(list)
    for row in crosswalk:
        soc = row['soc_code']
        isco = row['isco08_code']
        if soc in soc_probs:
            isco_groups[isco].append({
                'soc_code': soc,
                'probability': soc_probs[soc],
                'mapping_type': row['mapping_type'],
                'isco_title': row['isco08_title'],
            })
    
    results = []
    for isco_code, entries in sorted(isco_groups.items()):
        probs = [e['probability'] for e in entries]
        soc_codes = sorted(set(e['soc_code'] for e in entries))
        
        # Mapping confidence based on mapping types
        types = [e['mapping_type'] for e in entries]
        if all(t == 'direct' for t in types):
            confidence = 'high'
        elif any(t == 'direct' for t in types):
            confidence = 'medium'
        else:
            confidence = 'low'
        
        results.append({
            'isco08_code': isco_code,
            'isco08_title': entries[0]['isco_title'],
            'automation_probability': round(sum(probs) / len(probs), 4),
            'source_soc_codes': '|'.join(soc_codes),
            'mapping_confidence': confidence,
        })
    
    return results


def main():
    print("=" * 60)
    print("SOC ↔ ISCO-08 Crosswalk Builder")
    print("=" * 60)
    
    # Load source data
    print("\n[1/5] Loading source data...")
    frey_osborne = load_frey_osborne()
    print(f"  Frey-Osborne: {len(frey_osborne)} occupations")
    
    isco = load_isco08()
    isco_detailed = {k: v for k, v in isco.items() if v['level'] == 4}
    print(f"  ISCO-08: {len(isco)} total entries, {len(isco_detailed)} detailed (level 4)")
    
    # Try BLS download first
    print("\n[2/5] Attempting BLS crosswalk download...")
    bls_data = try_download_bls_crosswalk()
    
    if bls_data:
        print(f"  Using official BLS crosswalk ({len(bls_data)} rows)")
        # TODO: Parse BLS format into our standard format
        crosswalk = bls_data
    else:
        print("  BLS download failed — building algorithmic crosswalk")
        crosswalk = build_algorithmic_crosswalk(frey_osborne, isco)
    
    print(f"\n[3/5] Crosswalk built: {len(crosswalk)} mappings")
    
    # Write SOC↔ISCO crosswalk
    out_crosswalk = DATA_DIR / "soc_isco_crosswalk.csv"
    with open(out_crosswalk, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['soc_code', 'soc_title', 'isco08_code', 'isco08_title', 'mapping_type'])
        writer.writeheader()
        writer.writerows(crosswalk)
    print(f"  ✓ Written: {out_crosswalk} ({len(crosswalk)} rows)")
    
    # Build Frey-Osborne → ISCO mapping
    print("\n[4/5] Mapping Frey-Osborne to ISCO-08...")
    fo_isco = build_frey_osborne_isco(crosswalk, frey_osborne)
    
    out_fo = DATA_DIR / "frey_osborne_isco.csv"
    with open(out_fo, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['isco08_code', 'isco08_title', 'automation_probability', 'source_soc_codes', 'mapping_confidence'])
        writer.writeheader()
        writer.writerows(fo_isco)
    print(f"  ✓ Written: {out_fo} ({len(fo_isco)} ISCO codes with automation scores)")
    
    # Summary statistics
    print("\n[5/5] Computing summary statistics...")
    unique_soc = set(r['soc_code'] for r in crosswalk)
    unique_isco = set(r['isco08_code'] for r in crosswalk)
    all_soc = set(fo['soc_code'] for fo in frey_osborne)
    unmapped_soc = all_soc - unique_soc
    
    mapping_types = defaultdict(int)
    for r in crosswalk:
        mapping_types[r['mapping_type']] += 1
    
    confidence_dist = defaultdict(int)
    for r in fo_isco:
        confidence_dist[r['mapping_confidence']] += 1
    
    summary = {
        'total_soc_codes_in_source': len(all_soc),
        'total_soc_codes_mapped': len(unique_soc),
        'total_isco_codes_mapped': len(unique_isco),
        'total_crosswalk_rows': len(crosswalk),
        'unmapped_soc_count': len(unmapped_soc),
        'unmapped_soc_codes': sorted(list(unmapped_soc)),
        'coverage_percentage': round(len(unique_soc) / len(all_soc) * 100, 1),
        'mapping_quality_distribution': dict(mapping_types),
        'frey_osborne_isco_count': len(fo_isco),
        'frey_osborne_confidence_distribution': dict(confidence_dist),
        'method': 'algorithmic_title_matching' if not bls_data else 'bls_official',
    }
    
    out_summary = DATA_DIR / "crosswalk_summary.json"
    with open(out_summary, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"  ✓ Written: {out_summary}")
    
    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  SOC codes mapped:     {summary['total_soc_codes_mapped']}/{summary['total_soc_codes_in_source']} ({summary['coverage_percentage']}%)")
    print(f"  ISCO codes mapped:    {summary['total_isco_codes_mapped']}")
    print(f"  Crosswalk rows:       {summary['total_crosswalk_rows']}")
    print(f"  Unmapped SOC codes:   {summary['unmapped_soc_count']}")
    print(f"  Mapping quality:      {dict(mapping_types)}")
    print(f"  F&O → ISCO codes:     {summary['frey_osborne_isco_count']}")
    print(f"  F&O confidence:       {dict(confidence_dist)}")
    print(f"  Method:               {summary['method']}")
    print("=" * 60)
    
    # Validation
    ok = True
    if len(crosswalk) < 600:
        print(f"  ⚠ WARNING: crosswalk has {len(crosswalk)} rows (target: 600+)")
        ok = False
    if len(fo_isco) < 200:
        print(f"  ⚠ WARNING: F&O ISCO has {len(fo_isco)} codes (target: 200+)")
        ok = False
    if ok:
        print("  ✓ All acceptance criteria met!")
    
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
